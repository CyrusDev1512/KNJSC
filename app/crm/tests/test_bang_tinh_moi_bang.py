"""Bảng tính cho mọi bảng — `docs/04` mục 11, ADR-010.

Lưới không còn là của riêng bảng vận đơn: bảng nào trong phạm vi quyền cũng
mở được ở `/bang-tinh/<mã>/`. Mỗi bài phân quyền kiểm cả hai chiều.
"""
from datetime import date, timedelta
from io import BytesIO

import pytest
from django.core.exceptions import ValidationError
from django.test import override_settings

from core.constants import GRID_SPARE_ROWS, AuditAction
from core.models import AuditLog
from forms_builder.meaning import FieldType, Meaning
from forms_builder.models import ColumnDef, ComputeOp, DataRecord, GrantAction, TableDef
from forms_builder.services import grant_service, record_service, table_service
from orders.models import Product, ProductGroup
from orders.services import dispatch_service

pytestmark = pytest.mark.django_db

SUA_DUOC = override_settings(GRID_ONLY_TABLES=set())    # đúng cấu hình dịch vụ bangtinh


@pytest.fixture
def bang_sale(departments, nguoi_dung):
    """Bảng của bộ phận Sale: cột ngày, khách, sản phẩm mang nhãn, một cột tính sẵn."""
    bang = TableDef.objects.create(
        name="Đơn hàng Sale", code="don_sale",
        department=departments["sale"], created_by=nguoi_dung["manager_sale"],
    )
    cot = [
        ("Ngày", "ngay", FieldType.DATE, Meaning.DATE),
        ("Khách hàng", "khach", FieldType.TEXT, Meaning.CUSTOMER),
        ("Sản phẩm", "san_pham", FieldType.TEXT, Meaning.PRODUCT),
        ("Doanh thu", "doanh_thu", FieldType.MONEY, Meaning.REVENUE),
        ("Số lượng", "so_luong", FieldType.INTEGER, ""),
    ]
    for i, (ten, ma, kieu, nhan) in enumerate(cot):
        ColumnDef.objects.create(table=bang, name=ten, code=ma, field_type=kieu, meaning=nhan, order=i)
    ColumnDef.objects.create(
        table=bang, name="Giá đơn vị", code="gia_dv", field_type=FieldType.MONEY,
        order=5, is_computed=True, compute_op=ComputeOp.DIVIDE,
        compute_left="doanh_thu", compute_right="so_luong", compute_decimals=2,
    )
    return bang


@pytest.fixture
def bang_mkt(departments, nguoi_dung):
    return TableDef.objects.create(
        name="Báo cáo Marketing", code="bc_mkt",
        department=departments["mkt"], created_by=nguoi_dung["manager_mkt"],
    )


@pytest.fixture
def san_pham(db):
    nhom = ProductGroup.objects.create(name="Mỹ phẩm")
    return {
        "cream": Product.objects.create(name="Retinol Cream", code="retinol-cream", group=nhom),
        "serum": Product.objects.create(name="Retinol Serum", code="retinol-serum", group=nhom),
    }


@pytest.fixture
def bang_vd(departments, nguoi_dung, san_pham):
    return dispatch_service.ensure_waybill_table(actor=nguoi_dung["admin"])


def _dong(bang, nguoi, **gia_tri):
    return record_service.create_record(bang, gia_tri, actor=nguoi)


def _so_dong(client, duong):
    kq = client.get(duong)
    assert kq.status_code == 200, duong
    return kq.context["page_obj"].paginator.count


# ══ Mọi bảng trong phạm vi — AC-11.12 ══════════════════════════════

def test_moi_bang_trong_pham_vi_mo_duoc_o_bang_tinh(client, bang_sale, bang_mkt, nguoi_dung,
                                                    django_assert_max_num_queries):
    """AC-11.12 — Bảng nào trong phạm vi quyền cũng mở được ở /bang-tinh/<mã>/, ngoài phạm vi thì 404; /bang-tinh/ chọn bảng mặc định; thanh công cụ hiện nút theo quyền"""
    _dong(bang_sale, nguoi_dung["staff_sale_1"], ngay="2026-08-01", khach="A", doanh_thu="100", so_luong="2")

    for ma in ("staff_sale_1", "leader_sale_1", "manager_sale", "admin"):
        client.force_login(nguoi_dung[ma])
        kq = client.get("/bang-tinh/don_sale/")
        assert kq.status_code == 200 and kq.context["bang"].code == "don_sale", ma
        assert "don_sale" in [b.code for b in kq.context["cac_bang"]]
        # bảng không thuộc phạm vi thì 404 — quản trị viên thì thấy tất
        mong = 200 if ma == "admin" else 404
        assert client.get("/bang-tinh/bc_mkt/").status_code == mong, ma
        if ma != "admin":
            assert "bc_mkt" not in [b.code for b in kq.context["cac_bang"]]

    for ma in ("staff_vd", "staff_mkt"):
        client.force_login(nguoi_dung[ma])
        assert client.get("/bang-tinh/don_sale/").status_code == 404, ma
        assert client.get("/bang-tinh/don_sale/loc/khach/").status_code == 404
        assert client.post("/bang-tinh/don_sale/dong-moi/", {"khach": "x"}).status_code == 404

    # Bảng mặc định: không thấy bảng vận đơn thì lấy bảng đầu tiên trong phạm vi
    client.force_login(nguoi_dung["manager_sale"])
    assert client.get("/bang-tinh/").context["bang"].code == "don_sale"
    # Không có bảng nào trong phạm vi thì 404 kèm lời giải thích, không trang trắng
    client.force_login(nguoi_dung["staff_vd"])
    assert client.get("/bang-tinh/").status_code == 404

    # Thanh công cụ theo quyền: Manager thêm cột và nhập tệp, Staff thì không; ai cũng xuất được
    client.force_login(nguoi_dung["manager_sale"])
    html = client.get("/bang-tinh/don_sale/").content.decode()
    assert "Thêm cột" in html and "Nhập Excel" in html and "Tải Excel" in html
    client.force_login(nguoi_dung["staff_sale_1"])
    html = client.get("/bang-tinh/don_sale/").content.decode()
    assert "Thêm cột" not in html and "Nhập Excel" not in html and "Tải Excel" in html
    assert "Thêm dòng" in html                       # cùng bộ phận thì thêm dòng được
    assert 'class="bt-ben"' in html and 'class="bt-cong-cu"' in html

    # Ngân sách truy vấn: thanh bên thêm không quá ba lệnh so với lưới cũ (K24: 12)
    client.get("/bang-tinh/don_sale/")
    with django_assert_max_num_queries(14):
        client.get("/bang-tinh/don_sale/")

    client.logout()
    kq = client.get("/bang-tinh/don_sale/")
    assert kq.status_code == 302 and "/dang-nhap/" in kq["Location"]


# ══ Thanh lọc bên trái — AC-11.13 ══════════════════════════════════

def test_thanh_ben_chon_nhanh_khoang_ngay_san_pham(client, bang_sale, bang_vd, san_pham, nguoi_dung):
    """AC-11.13 — Chọn nhanh và khoảng ngày viết vào bộ lọc cột Ngày, sản phẩm lọc "có một trong", xuất Excel ra đúng số dòng của lưới đang lọc"""
    hom_nay = date.today()
    hom_qua = hom_nay - timedelta(days=1)
    xa = hom_nay - timedelta(days=40)
    st = nguoi_dung["staff_sale_1"]
    _dong(bang_sale, st, ngay=hom_nay.isoformat(), khach="A", san_pham="Kem", doanh_thu="10", so_luong="1")
    _dong(bang_sale, st, ngay=hom_qua.isoformat(), khach="B", san_pham="Serum", doanh_thu="10", so_luong="1")
    _dong(bang_sale, st, ngay=xa.isoformat(), khach="C", san_pham="Kem", doanh_thu="10", so_luong="1")
    client.force_login(st)

    kq = client.get("/bang-tinh/don_sale/")
    ben = kq.context["ben"]
    assert ben["cot_ngay"].code == "ngay" and ben["san_pham"]["kind"] == "gia_tri"
    chon_nhanh = {ma: (qs, bat) for ma, _, qs, bat, _ in ben["chon_nhanh"]}
    qs_hom_qua, _ = chon_nhanh["hom_qua"]
    assert f"f_ngay__lon_bang={hom_qua.isoformat()}" in qs_hom_qua
    assert f"f_ngay__nho_bang={hom_qua.isoformat()}" in qs_hom_qua
    assert _so_dong(client, "/bang-tinh/don_sale/?" + qs_hom_qua) == 1
    assert _so_dong(client, "/bang-tinh/don_sale/?" + chon_nhanh["7_ngay"][0]) == 2
    # mốc đang bật thì chip đánh dấu, và chọn nhanh giữ các bộ lọc khác
    kq = client.get("/bang-tinh/don_sale/?" + qs_hom_qua + "&f_khach__chua=B")
    assert dict((ma, bat) for ma, _, _, bat, _ in kq.context["ben"]["chon_nhanh"])["hom_qua"] is True
    assert "f_khach__chua" in dict((ma, qs) for ma, _, qs, _, _ in kq.context["ben"]["chon_nhanh"])["hom_nay"]
    # từ ngày / đến ngày gõ tay
    assert _so_dong(client, f"/bang-tinh/don_sale/?f_ngay__lon_bang={xa}&f_ngay__nho_bang={hom_qua}") == 2

    # Sản phẩm: chọn một hoặc nhiều — "có một trong"
    assert ben["san_pham"]["param"] == "f_san_pham__trong"
    assert {gt for gt, *_ in ben["san_pham"]["items"]} == {"Kem", "Serum"}
    assert _so_dong(client, "/bang-tinh/don_sale/?f_san_pham__trong=Kem") == 2
    assert _so_dong(client, "/bang-tinh/don_sale/?f_san_pham__trong=Kem&f_san_pham__trong=Serum") == 3

    # Bảng vận đơn: mỗi sản phẩm một cột số lượng → lọc bằng sp=<mã cột>
    vd = nguoi_dung["staff_vd"]
    _dong(bang_vd, vd, ma_don="D1", ten_khach="X", so_dien_thoai="0911", sl_retinol_cream=2)
    _dong(bang_vd, vd, ma_don="D2", ten_khach="Y", so_dien_thoai="0911", sl_retinol_serum=1)
    _dong(bang_vd, vd, ma_don="D3", ten_khach="Z", so_dien_thoai="0922")
    client.force_login(vd)
    ben = client.get("/bang-tinh/van_don/").context["ben"]
    assert ben["san_pham"]["kind"] == "cot_sl" and ben["san_pham"]["param"] == "sp"
    assert {gt for gt, *_ in ben["san_pham"]["items"]} == {"sl_retinol_cream", "sl_retinol_serum"}
    assert _so_dong(client, "/bang-tinh/van_don/?sp=sl_retinol_cream") == 1
    assert _so_dong(client, "/bang-tinh/van_don/?sp=sl_retinol_cream&sp=sl_retinol_serum") == 2
    assert _so_dong(client, "/bang-tinh/van_don/?sp=khong_co") == 3       # mã lạ bị bỏ qua
    assert _so_dong(client, "/bang-tinh/van_don/?sp=sl_retinol_cream&trung=1") == 1

    # Xuất Excel đúng lưới đang hiện — kể cả hai bộ lọc riêng của lưới (ADR-002)
    from openpyxl import load_workbook

    kq = client.get("/bang-tinh/van_don/xuat/?sp=sl_retinol_cream&sp=sl_retinol_serum&trung=1")
    assert kq.status_code == 200
    ws = load_workbook(BytesIO(kq.content)).active
    assert ws.max_row - 1 == 2                      # trừ hàng tiêu đề
    kq = client.get("/bang-tinh/van_don/xuat/?sp=sl_retinol_serum")
    assert load_workbook(BytesIO(kq.content)).active.max_row - 1 == 1


# ══ Dòng trống cuối lưới — AC-11.14 ════════════════════════════════

def test_dong_trong_sinh_dong_that_khi_nhap_o_dau(client, bang_sale, bang_vd, departments, nguoi_dung):
    """AC-11.14 — Lưới thừa dòng trống cho người có quyền thêm; gõ vào rồi gửi là thành bản ghi thật đúng bộ phận; lỗi thì 400 kèm lý do và giữ giá trị; không quyền thì không có dòng trống và POST bị 403 có ghi nhật ký"""
    st = nguoi_dung["staff_sale_1"]
    client.force_login(st)
    html = client.get("/bang-tinh/don_sale/").content.decode()
    assert html.count('class="dong-moi"') == GRID_SPARE_ROWS
    assert 'name="gia_dv"' not in html            # cột tính sẵn không có ô nhập

    # Gửi dòng: thành bản ghi thật, trả về dòng thật + một dòng trống mới
    kq = client.post("/bang-tinh/don_sale/dong-moi/", {
        "ngay": "2026-08-01", "khach": "Khách mới", "doanh_thu": "100", "so_luong": "2",
    }, HTTP_HX_CURRENT_URL="http://testserver/bang-tinh/don_sale/?f_khach__chua=K")
    assert kq.status_code == 200
    moi = DataRecord.objects.get(table=bang_sale)
    html = kq.content.decode()
    assert f'data-dong="{moi.pk}"' in html and html.count('class="dong-moi"') == 1
    assert moi.created_by == st and moi.department == departments["sale"]
    assert moi.val_customer == "Khách mới" and moi.data["gia_dv"] == "50.00"

    # Dòng trống hoàn toàn → 400; ngày sai → 400, giữ giá trị đã gõ, tô ô lỗi
    assert client.post("/bang-tinh/don_sale/dong-moi/", {}).status_code == 400
    kq = client.post("/bang-tinh/don_sale/dong-moi/", {"ngay": "abc", "khach": "Lỗi"})
    assert kq.status_code == 400
    html = kq.content.decode()
    assert "o-loi" in html and 'value="abc"' in html and 'value="Lỗi"' in html
    assert DataRecord.objects.filter(table=bang_sale).count() == 1
    # Thiếu cột bắt buộc
    bang_sale.columns.filter(code="khach").update(required=True)
    kq = client.post("/bang-tinh/don_sale/dong-moi/", {"ngay": "2026-08-02"})
    assert kq.status_code == 400 and "bắt buộc" in kq.content.decode()

    # Người được cấp quyền XEM từ bộ phận khác: thấy lưới, không có dòng trống, POST bị 403
    mkt = nguoi_dung["staff_mkt"]
    grant_service.grant(table=bang_sale, user=mkt, action=GrantAction.VIEW, actor=nguoi_dung["manager_sale"])
    client.force_login(mkt)
    kq = client.get("/bang-tinh/don_sale/")
    assert kq.status_code == 200 and "dong-moi" not in kq.content.decode()
    truoc = AuditLog.objects.filter(action=AuditAction.DENIED).count()
    assert client.post("/bang-tinh/don_sale/dong-moi/", {"khach": "Lén"}).status_code == 403
    assert AuditLog.objects.filter(action=AuditAction.DENIED).count() == truoc + 1
    assert DataRecord.objects.filter(table=bang_sale).count() == 1
    # Cấp thêm quyền SỬA thì thêm được
    grant_service.grant(table=bang_sale, user=mkt, action=GrantAction.EDIT, actor=nguoi_dung["manager_sale"])
    assert client.post("/bang-tinh/don_sale/dong-moi/", {"ngay": "2026-08-03", "khach": "MKT thêm"}).status_code == 200

    # Bảng vận đơn: chỉ xem ở dịch vụ chính → không dòng trống; ở dịch vụ Bảng tính thì có
    vd = nguoi_dung["staff_vd"]
    client.force_login(vd)
    assert "dong-moi" not in client.get("/bang-tinh/van_don/").content.decode()
    assert client.post("/bang-tinh/van_don/dong-moi/", {"ma_don": "D9"}).status_code == 403
    with SUA_DUOC:
        assert client.get("/bang-tinh/van_don/").content.decode().count('class="dong-moi"') == GRID_SPARE_ROWS
        kq = client.post("/bang-tinh/van_don/dong-moi/", {"ma_don": "D9", "ten_khach": "Mới", "so_dien_thoai": "0911"})
        assert kq.status_code == 200
    assert DataRecord.objects.filter(table=bang_vd, val_phone="0911").count() == 1


# ══ Cột khoá — AC-11.16 ════════════════════════════════════════════

def test_cot_khoa_mot_cot_moi_bang_va_loc_theo_o(client, bang_sale, bang_vd, nguoi_dung):
    """AC-11.16 — Mỗi bảng một cột khoá do Manager đặt; ô cột khoá có liên kết lọc theo giá trị, theo liên kết ra đúng dòng; cột tính sẵn không làm khoá được"""
    ql = nguoi_dung["manager_sale"]
    cot_ma = table_service.add_column(
        bang_sale, actor=ql, name="Mã đơn", code="ma", field_type=FieldType.TEXT, is_key=True,
    )
    assert cot_ma.is_key
    with pytest.raises(ValidationError):
        table_service.add_column(bang_sale, actor=ql, name="Mã 2", code="ma2", field_type=FieldType.TEXT, is_key=True)
    with pytest.raises(ValidationError):
        table_service.update_column(bang_sale.columns.get(code="gia_dv"), {"is_key": True}, actor=ql)
    with pytest.raises(ValidationError):
        table_service.update_column(bang_sale.columns.get(code="ngay"), {"is_key": True}, actor=ql)
    assert bang_sale.columns.filter(is_key=True).count() == 1
    # Bảng vận đơn: Mã đơn là cột khoá sẵn
    assert bang_vd.columns.get(code="ma_don").is_key

    st = nguoi_dung["staff_sale_1"]
    _dong(bang_sale, st, ma="A1", ngay="2026-08-01", khach="Một", doanh_thu="10", so_luong="1")
    _dong(bang_sale, st, ma="A2", ngay="2026-08-01", khach="Hai", doanh_thu="10", so_luong="1")
    client.force_login(st)
    kq = client.get("/bang-tinh/don_sale/?sap=ma")
    assert kq.context["cot_khoa"].code == "ma"
    html = kq.content.decode()
    assert html.count('class="o-khoa-loc"') == 2
    assert "?sap=ma&amp;f_ma=A1" in html or "?sap=ma&f_ma=A1" in html   # cộng dồn tham số đang bật
    assert 'class="th-khoa' in html or "th-khoa" in html
    assert _so_dong(client, "/bang-tinh/don_sale/?f_ma=A1") == 1
    assert _so_dong(client, "/bang-tinh/don_sale/?f_ma=A1&f_khach__chua=Hai") == 0

    # Sửa cột trên màn hình Sửa cột: Manager đặt được, Leader bị 403
    client.force_login(ql)
    kq = client.post(f"/bang/don_sale/cot/?cot={cot_ma.pk}", {
        "name": "Mã đơn", "code": "ma", "field_type": "text", "meaning": "", "order": 6,
        "compute_decimals": 2, "compute_op": "", "compute_left": "", "compute_right": "",
    })
    assert kq.status_code == 302
    assert not bang_sale.columns.get(code="ma").is_key              # bỏ tích thì hết khoá
    kq = client.post(f"/bang/don_sale/cot/?cot={cot_ma.pk}", {
        "name": "Mã đơn", "code": "ma", "field_type": "text", "meaning": "", "order": 6, "is_key": "on",
        "compute_decimals": 2, "compute_op": "", "compute_left": "", "compute_right": "",
    })
    assert kq.status_code == 302 and bang_sale.columns.get(code="ma").is_key
    assert "Khoá" in client.get("/bang/don_sale/cot/").content.decode()
    client.force_login(nguoi_dung["leader_sale_1"])
    assert client.post(f"/bang/don_sale/cot/?cot={cot_ma.pk}", {"name": "x"}).status_code == 403
