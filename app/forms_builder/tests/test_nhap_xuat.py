"""Nhập và xuất tệp Excel trên bảng dữ liệu — FR-7.5 → FR-7.7, đi qua HTTP.

`test_nhap_hang_loat.py` kiểm tầng dịch vụ ghi hàng loạt; tệp này kiểm trọn
luồng người dùng thật chạm vào: chọn tệp → xem trước → xác nhận → tác vụ nền
→ trang tiến độ, và chiều ngược lại: xuất đúng thứ đang hiện rồi nhập lại
được. Mỗi bài phân quyền kiểm **cả hai chiều**.
"""
import time
from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import override_settings
from openpyxl import load_workbook

from core import excel
from core.constants import (
    EXPORT_SYNC_MAX_ROWS, IMPORT_MAX_ROWS, IMPORT_PERF_ROWS, IMPORT_PERF_SECONDS,
    UPLOAD_MAX_BYTES, AuditAction, JobStatus,
)
from core.models import AuditLog, BackgroundJob
from forms_builder.meaning import FieldType, Meaning
from forms_builder.models import ColumnDef, DataRecord, GrantAction, TableDef
from forms_builder.services import export_service, grant_service, import_service, record_service

pytestmark = pytest.mark.django_db

XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture
def bang_sale(departments, nguoi_dung):
    bang = TableDef.objects.create(
        name="Đơn hàng Sale", code="don_sale",
        department=departments["sale"], created_by=nguoi_dung["manager_sale"],
    )
    cot = [
        ("Ngày", "ngay", FieldType.DATE, Meaning.DATE, False),
        ("Khách hàng", "khach", FieldType.TEXT, Meaning.CUSTOMER, True),
        ("Doanh thu", "doanh_thu", FieldType.MONEY, Meaning.REVENUE, False),
        ("Số lượng", "so_luong", FieldType.INTEGER, "", False),
    ]
    for i, (ten, ma, kieu, nhan, bat_buoc) in enumerate(cot):
        ColumnDef.objects.create(
            table=bang, name=ten, code=ma, field_type=kieu, meaning=nhan,
            required=bat_buoc, order=i,
        )
    return bang


def _tep_xlsx(headers, rows, ten="don.xlsx"):
    """Tệp Excel tải lên, dựng từ bảng giá trị."""
    wb = excel.write_table(headers, rows)
    dem = BytesIO()
    wb.save(dem)
    return SimpleUploadedFile(ten, dem.getvalue(), content_type=XLSX)


def _tep_don(rows, ten="don.xlsx"):
    return _tep_xlsx(["Ngày", "Khách hàng", "Doanh thu", "Số lượng"], rows, ten)


def _nhap_tron_luong(client, code, tep):
    """Đi trọn ba bước qua HTTP, trả về tác vụ sau khi worker (eager) chạy xong."""
    kq = client.post(f"/bang/{code}/nhap/", {"tep": tep})
    assert kq.status_code == 302, kq.content.decode()[:500]
    duong_dan_xem_truoc = kq["Location"]
    assert client.get(duong_dan_xem_truoc).status_code == 200
    kq = client.post(duong_dan_xem_truoc + "xac-nhan/")
    assert kq.status_code == 302
    pk = int(duong_dan_xem_truoc.rstrip("/").rsplit("/", 1)[-1])
    return BackgroundJob.objects.get(pk=pk)


def _doc_xlsx(noi_dung):
    ws = load_workbook(BytesIO(noi_dung), read_only=True).worksheets[0]
    return [list(r) for r in ws.iter_rows(values_only=True)]


# ══ Luồng nhập ═════════════════════════════════════════════════════

def test_nhap_tep_co_dong_loi_van_nhap_dong_hop_le(client, bang_sale, nguoi_dung):
    """AC-7.6 — Tệp 5 dòng có 2 dòng lỗi thì 3 dòng vào, 2 dòng lỗi được liệt kê theo số hàng Excel"""
    client.force_login(nguoi_dung["manager_sale"])
    tep = _tep_don([
        [date(2026, 8, 1), "A", 100, 1],
        [date(2026, 8, 2), "B", "abc", 1],       # tiền không đọc được → hàng 3
        [date(2026, 8, 3), "C", 300, 3],
        [date(2026, 8, 4), None, 400, 4],        # thiếu khách hàng → hàng 5
        [date(2026, 8, 5), "E", 500, 5],
    ])
    job = _nhap_tron_luong(client, "don_sale", tep)
    assert job.status == JobStatus.DONE
    assert job.summary["created"] == 3
    assert [so for so, _ in job.summary["errors"]] == [3, 5]
    assert DataRecord.objects.filter(table=bang_sale).count() == 3

    trang = client.get(f"/tac-vu/{job.pk}/").content.decode()
    assert '<td class="ma-o">3</td>' in trang and '<td class="ma-o">5</td>' in trang
    # Mảnh tiến độ: xong rồi thì không còn hỏi lại nữa
    manh = client.get(f"/tac-vu/{job.pk}/tien-do/").content.decode()
    assert "hx-trigger" not in manh


def test_xem_truoc_chua_ghi_gi(client, bang_sale, nguoi_dung):
    """FR-7.5 — Chưa xác nhận thì chưa có dòng nào được ghi vào bảng"""
    client.force_login(nguoi_dung["manager_sale"])
    kq = client.post("/bang/don_sale/nhap/", {"tep": _tep_don([[date(2026, 8, 1), "A", 100, 1]])})
    assert kq.status_code == 302
    job = BackgroundJob.objects.latest("id")
    assert job.status == JobStatus.DRAFT and job.total == 1
    assert DataRecord.objects.filter(table=bang_sale).count() == 0
    noi_dung = client.get(kq["Location"]).content.decode()
    assert "Khách hàng" in noi_dung and "Xác nhận" in noi_dung


def test_xuat_roi_nhap_lai_khong_loi(client, bang_sale, nguoi_dung):
    """AC-7.7 — Tệp xuất ra nhập lại được, không dòng nào lỗi, giá trị giữ nguyên"""
    manager = nguoi_dung["manager_sale"]
    for i in range(1, 4):
        record_service.create_record(
            bang_sale, {"ngay": f"2026-08-0{i}", "khach": f"K{i}",
                        "doanh_thu": f"{i}000.50", "so_luong": i},
            actor=manager,
        )
    client.force_login(manager)
    kq = client.get("/bang/don_sale/xuat/")
    assert kq.status_code == 200 and kq["Content-Type"] == XLSX
    hang = _doc_xlsx(kq.content)
    assert hang[0] == ["Ngày", "Khách hàng", "Doanh thu", "Số lượng"]
    assert len(hang) == 4

    job = _nhap_tron_luong(
        client, "don_sale", SimpleUploadedFile("xuat.xlsx", kq.content, content_type=XLSX),
    )
    assert job.status == JobStatus.DONE
    assert job.summary["created"] == 3 and job.summary["errors"] == []
    moi = DataRecord.objects.filter(table=bang_sale).order_by("-id")[:3]
    assert sorted(Decimal(r.data["doanh_thu"]) for r in moi) == [Decimal("1000.50"), Decimal("2000.50"), Decimal("3000.50")]
    assert {r.data["ngay"] for r in moi} == {"2026-08-01", "2026-08-02", "2026-08-03"}


def test_tep_qua_10mb_bi_tu_choi(client, bang_sale, nguoi_dung):
    """AC-7.8 — Tệp lớn hơn 10 MB bị từ chối, thông báo rõ, không có tác vụ nào được tạo"""
    client.force_login(nguoi_dung["manager_sale"])
    truoc = BackgroundJob.objects.count()
    tep = SimpleUploadedFile("to.xlsx", b"PK\x03\x04" + b"\0" * UPLOAD_MAX_BYTES, content_type=XLSX)
    kq = client.post("/bang/don_sale/nhap/", {"tep": tep}, follow=True)
    assert kq.status_code == 200
    assert "vượt giới hạn" in kq.content.decode()
    assert BackgroundJob.objects.count() == truoc


@pytest.mark.parametrize("ten, noi_dung", [
    ("virus.xlsx", b"MZ\x90\x00" + b"\0" * 64),                  # tệp .exe đổi đuôi
    ("anh.csv", b"\x89PNG\r\n\x1a\n" + b"\0" * 64),              # ảnh PNG đổi đuôi
    ("bang.xlsx", b"ngay,khach\n2026-08-01,A\n"),               # CSV nhưng khai .xlsx
])
def test_tep_sai_loai_bi_tu_choi(client, bang_sale, nguoi_dung, ten, noi_dung):
    """AC-7.9 — Tệp không phải Excel/CSV, hoặc đuôi khai một đằng nội dung một nẻo, bị từ chối"""
    client.force_login(nguoi_dung["manager_sale"])
    kq = client.post("/bang/don_sale/nhap/", {"tep": SimpleUploadedFile(ten, noi_dung)}, follow=True)
    assert kq.status_code == 200
    assert "không" in kq.content.decode().lower()
    assert not BackgroundJob.objects.exists()


def test_tep_hon_5000_dong_bi_tu_choi(client, bang_sale, nguoi_dung):
    """NFR-13 — Tệp vượt trần 5.000 dòng bị từ chối trước khi ghi gì"""
    client.force_login(nguoi_dung["manager_sale"])
    tep = _tep_don([[date(2026, 8, 1), f"K{i}", 100, 1] for i in range(IMPORT_MAX_ROWS + 1)])
    kq = client.post("/bang/don_sale/nhap/", {"tep": tep}, follow=True)
    assert "5.000" in kq.content.decode()
    assert DataRecord.objects.filter(table=bang_sale).count() == 0
    assert not BackgroundJob.objects.exists()


def test_nhap_tep_van_don_that_do_tieu_de_hang_2(departments, nguoi_dung):
    """FR-7.5 — Tệp vận đơn thật (ẩn danh): tiêu đề ở hàng 2, hàng công thức bị bỏ, điện thoại là chuỗi

    Tệp `docs/tham-khao/vandon-mau.xlsx` giữ nguyên mọi "bẫy" của tệp gốc.
    Bài này chỉ dừng ở bước xem trước; nhập trọn tệp là AC-11.9 ở Giai đoạn 7C
    khi bảng vận đơn có đủ cột.
    """
    from pathlib import Path

    from orders.services.dispatch_service import ensure_waybill_table

    bang = ensure_waybill_table(actor=nguoi_dung["admin"])
    duong_dan = Path(__file__).resolve().parents[3] / "docs" / "tham-khao" / "vandon-mau.xlsx"
    with open(duong_dan, "rb") as f:
        tep = SimpleUploadedFile("vandon-mau.xlsx", f.read(), content_type=XLSX)
    job = import_service.prepare(bang, tep, actor=nguoi_dung["admin"])
    assert job.status == JobStatus.DRAFT
    assert job.summary["header_row"] == 2
    assert job.total == 221
    khop = {m["code"] for m in job.summary["mapping"]}
    assert {"ten_khach", "so_dien_thoai", "ngay", "gia_tien", "trang_thai_vc"} <= khop
    assert job.summary["ignored"], "cột dư trong tệp phải được báo, không im lặng"
    # Điện thoại Excel lưu dạng số thực phải về chuỗi số, không ".0"
    cot = [m["code"] for m in job.summary["mapping"]]
    dien_thoai = job.summary["sample"][0][cot.index("so_dien_thoai")]
    assert dien_thoai and not dien_thoai.endswith(".0")


# ══ Luồng xuất ═════════════════════════════════════════════════════

def test_xuat_kem_bo_loc_va_sap_xep(client, bang_sale, nguoi_dung):
    """FR-7.6 — Xuất đúng thứ đang hiện: bộ lọc và thứ tự sắp xếp của màn hình đi theo tệp"""
    manager = nguoi_dung["manager_sale"]
    for ngay, khach in [("2026-08-01", "A"), ("2026-08-03", "A"), ("2026-08-02", "A"),
                        ("2026-08-04", "B")]:
        record_service.create_record(
            bang_sale, {"ngay": ngay, "khach": khach, "doanh_thu": "1", "so_luong": 1},
            actor=manager,
        )
    client.force_login(manager)
    kq = client.get("/bang/don_sale/xuat/?f_khach=A&sap=ngay&chieu=giam")
    hang = _doc_xlsx(kq.content)[1:]
    assert [h[1] for h in hang] == ["A", "A", "A"]
    assert [h[0].date() for h in hang] == [date(2026, 8, 3), date(2026, 8, 2), date(2026, 8, 1)]

    nhat_ky = AuditLog.objects.filter(action=AuditAction.EXPORT).latest("created_at")
    assert "3 dòng" in nhat_ky.detail and "lọc khach" in nhat_ky.detail


def test_xuat_lon_chay_nen_va_tai_duoc(client, bang_sale, nguoi_dung):
    """FR-7.6 — Trên 2.000 dòng thì xuất chạy nền: tải được ở Tác vụ nền, tệp bị dọn thì báo rõ"""
    manager = nguoi_dung["manager_sale"]
    record_service.create_records_bulk(
        bang_sale,
        [{"ngay": "2026-08-01", "khach": f"K{i}", "doanh_thu": "1", "so_luong": 1}
         for i in range(EXPORT_SYNC_MAX_ROWS + 1)],
        actor=manager,
    )
    client.force_login(manager)
    kq = client.get("/bang/don_sale/xuat/")
    assert kq.status_code == 302 and kq["Location"].startswith("/tac-vu/")
    job = BackgroundJob.objects.latest("id")
    assert job.status == JobStatus.DONE and job.result_path
    assert job.total == EXPORT_SYNC_MAX_ROWS + 1

    tai = client.get(f"/tac-vu/{job.pk}/tai/")
    assert tai.status_code == 200
    assert tai["Content-Disposition"].endswith('.xlsx"')
    noi_dung = b"".join(tai.streaming_content)
    assert len(_doc_xlsx(noi_dung)) == EXPORT_SYNC_MAX_ROWS + 2

    export_service.result_file(job).unlink()
    kq = client.get(f"/tac-vu/{job.pk}/tai/", follow=True)
    assert kq.status_code == 200 and "được dọn" in kq.content.decode()


def test_xuat_vuot_tran_bi_tu_choi(client, bang_sale, nguoi_dung):
    """NFR-14 — Kết quả vượt trần xuất thì từ chối với thông báo, không tạo tác vụ"""
    manager = nguoi_dung["manager_sale"]
    for i in range(3):
        record_service.create_record(
            bang_sale, {"ngay": "2026-08-01", "khach": f"K{i}", "doanh_thu": "1", "so_luong": 1},
            actor=manager,
        )
    client.force_login(manager)
    with override_settings(EXPORT_MAX_ROWS=2):
        kq = client.get("/bang/don_sale/xuat/", follow=True)
    assert "vượt giới hạn xuất" in kq.content.decode()
    assert not BackgroundJob.objects.exists()
    assert not AuditLog.objects.filter(action=AuditAction.EXPORT).exists()


# ══ Phân quyền — cả hai chiều ══════════════════════════════════════

def test_quyen_nhap_hai_chieu(client, bang_sale, nguoi_dung):
    """AC-3.6 — Nhân viên bị từ chối nhập (403 có ghi nhật ký), quản lý bộ phận sở hữu thì được"""
    client.force_login(nguoi_dung["staff_sale_1"])
    truoc = AuditLog.objects.filter(action=AuditAction.DENIED).count()
    assert client.get("/bang/don_sale/nhap/").status_code == 403
    assert client.post("/bang/don_sale/nhap/", {"tep": _tep_don([])}).status_code == 403
    assert AuditLog.objects.filter(action=AuditAction.DENIED).count() == truoc + 2

    client.force_login(nguoi_dung["leader_sale_1"])
    assert client.get("/bang/don_sale/nhap/").status_code == 403

    client.force_login(nguoi_dung["manager_mkt"])
    # Bộ phận khác không thấy bảng tồn tại — quy ước của màn hình bảng là 404
    assert client.get("/bang/don_sale/nhap/").status_code == 404

    client.force_login(nguoi_dung["manager_sale"])
    assert client.get("/bang/don_sale/nhap/").status_code == 200
    client.force_login(nguoi_dung["admin"])
    assert client.get("/bang/don_sale/nhap/").status_code == 200


def test_cap_quyen_sua_thi_nhap_duoc_cap_quyen_xem_thi_khong(client, bang_sale, nguoi_dung):
    """AC-3.6 — Cấp quyền Sửa trên bảng thì nhập được; chỉ cấp quyền Xem thì vẫn bị từ chối"""
    staff = nguoi_dung["staff_sale_1"]
    grant_service.grant(
        table=bang_sale, user=staff, action=GrantAction.VIEW, actor=nguoi_dung["manager_sale"],
    )
    client.force_login(staff)
    assert client.get("/bang/don_sale/nhap/").status_code == 403

    grant_service.grant(
        table=bang_sale, user=staff, action=GrantAction.EDIT, actor=nguoi_dung["manager_sale"],
    )
    client.force_login(staff)
    assert client.get("/bang/don_sale/nhap/").status_code == 200
    job = _nhap_tron_luong(client, "don_sale", _tep_don([[date(2026, 8, 1), "A", 100, 1]]))
    assert job.status == JobStatus.DONE and job.summary["created"] == 1


def test_tac_vu_nhap_cua_nguoi_khac_khong_thay(client, bang_sale, nguoi_dung):
    """AC-3.6 — Tác vụ nhập của người khác không mở được, dù có quyền nhập vào cùng bảng"""
    manager = nguoi_dung["manager_sale"]
    job = import_service.prepare(
        bang_sale, _tep_don([[date(2026, 8, 1), "A", 100, 1]]), actor=manager,
    )
    staff = nguoi_dung["staff_sale_1"]
    grant_service.grant(table=bang_sale, user=staff, action=GrantAction.EDIT, actor=manager)
    client.force_login(staff)
    assert client.get(f"/bang/don_sale/nhap/{job.pk}/").status_code == 404
    assert client.post(f"/bang/don_sale/nhap/{job.pk}/xac-nhan/").status_code == 404
    assert client.get(f"/tac-vu/{job.pk}/").status_code == 404
    assert client.get(f"/tac-vu/{job.pk}/tai/").status_code == 404
    job.refresh_from_db()
    assert job.status == JobStatus.DRAFT, "người khác không được kích hoạt tác vụ"

    # Chính chủ và Admin thì thấy
    client.force_login(manager)
    assert client.get(f"/bang/don_sale/nhap/{job.pk}/").status_code == 200
    client.force_login(nguoi_dung["admin"])
    assert client.get(f"/tac-vu/{job.pk}/").status_code == 200


def test_xuat_chi_ra_dong_trong_pham_vi(client, bang_sale, nguoi_dung):
    """AC-3.1 — Xuất tệp chỉ ra dòng người đó thấy được trên màn hình, không phải cả bảng"""
    for ma in ("staff_sale_1", "staff_sale_2", "manager_sale"):
        record_service.create_record(
            bang_sale, {"ngay": "2026-08-01", "khach": ma, "doanh_thu": "1", "so_luong": 1},
            actor=nguoi_dung[ma],
        )
    client.force_login(nguoi_dung["staff_sale_1"])
    hang = _doc_xlsx(client.get("/bang/don_sale/xuat/").content)[1:]
    assert [h[1] for h in hang] == ["staff_sale_1"]

    client.force_login(nguoi_dung["manager_sale"])
    assert len(_doc_xlsx(client.get("/bang/don_sale/xuat/").content)) == 4

    client.force_login(nguoi_dung["staff_mkt"])
    assert client.get("/bang/don_sale/xuat/").status_code == 404   # không thấy bảng


# ══ Hiệu năng ══════════════════════════════════════════════════════

@pytest.mark.cham
def test_nhap_2000_dong_duoi_60_giay(client, bang_sale, nguoi_dung):
    """AC-7.5 — Nhập tệp 2.000 dòng xong dưới 60 giây, đủ 2.000 dòng vào bảng"""
    client.force_login(nguoi_dung["manager_sale"])
    tep = _tep_don([[date(2026, 8, 1), f"K{i}", 100 + i, 1 + i % 5] for i in range(IMPORT_PERF_ROWS)])
    bat_dau = time.monotonic()
    job = _nhap_tron_luong(client, "don_sale", tep)
    mat = time.monotonic() - bat_dau
    assert job.status == JobStatus.DONE and job.summary["created"] == IMPORT_PERF_ROWS
    assert mat < IMPORT_PERF_SECONDS, f"nhập 2.000 dòng mất {mat:.1f}s"
    assert DataRecord.objects.filter(table=bang_sale).count() == IMPORT_PERF_ROWS
