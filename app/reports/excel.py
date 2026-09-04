"""Dựng tệp Excel cho báo cáo tổng hợp — FR-5.6.

Chỉ chứa cơ khí openpyxl, không truy vấn gì: nhận `SummaryResult` đã tính
xong và ghi ra Workbook. Số ghi dạng `Decimal` nguyên trạng, không đổi sang
float (BR-8) — Excel đọc được và số khớp tuyệt đối với màn hình.

Giai đoạn 7 sẽ có `core/excel.py` cho nhập và xuất bảng thô; tệp này chỉ
phục vụ báo cáo tổng hợp và có thể được gộp về đó sau.
"""
from openpyxl import Workbook
from openpyxl.styles import Font

from . import aggregations


def build_workbook(title, result, subtitle=""):
    """Một sheet: tiêu đề, dòng phụ, bảng số liệu, dòng cuối là tổng cộng."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Bao cao tong hop"

    dam = Font(bold=True)
    ws.append([title])
    ws["A1"].font = dam
    ws.append([subtitle])
    ws.append([])

    ws.append([result.group_label] + [c.label for c in result.columns])
    for o in ws[ws.max_row]:
        o.font = dam

    so_nhom = 0
    for item in result.rows.iterator():
        nhom, cells = aggregations.row_values(item, result)
        ws.append([aggregations.format_group(nhom, result)] + cells)
        so_nhom += 1

    ws.append([f"Tổng cộng · {so_nhom} {result.unit}"] + aggregations.total_values(result))
    for o in ws[ws.max_row]:
        o.font = dam

    return wb
