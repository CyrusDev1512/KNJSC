"""Màn hình Báo cáo tổng hợp qua HTTP — phạm vi quyền, bộ lọc, xuất Excel.

Mỗi bài phân quyền kiểm **cả hai chiều**: chiều thấy được và chiều không được
thấy — chỉ kiểm chiều được phép thì không phát hiện được rò rỉ dữ liệu.

Tab "Theo thị trường" đang hoãn (Q36, backlog N9) nên bài của nó ghi FR-5.1,
không ghi AC-5.1 — mã đó còn nằm trong danh sách hoãn của `test_truy_vet`.
"""
from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from core.constants import AuditAction
from core.models import AuditLog
from forms_builder.meaning import FieldType, Meaning
from forms_builder.models import ColumnDef, ComputeOp, TableDef
from forms_builder.services import record_service

pytestmark = pytest.mark.django_db

DUONG_DAN = "/bao-cao/tong-hop/"
DUONG_XUAT = "/bao-cao/tong-hop/xuat/"


@pytest.fixture
def bang_bc_sale(departments, nguoi_dung):
    """Bảng báo cáo của bộ phận Sale: bốn cột mang nhãn, một cột số trong
    JSON, một cột tính sẵn dạng chia."""
    bang = TableDef.objects.create(
        name="Báo cáo Sale", code="bc_sale",
        department=departments["sale"], created_by=nguoi_dung["manager_sale"],
    )
    cot = [
        ("Ngày", "ngay", FieldType.DATE, Meaning.DATE),
        ("Người bán", "nguoi_ban", FieldType.TEXT, Meaning.SELLER),
        ("Sản phẩm", "san_pham", FieldType.TEXT, Meaning.PRODUCT),
        ("Số đơn", "so_don", FieldType.INTEGER, ""),
        ("Doanh số", "doanh_so", FieldType.MONEY, Meaning.REVENUE),
    ]
    for i, (ten, ma, kieu, nhan) in enumerate(cot):
        ColumnDef.objects.create(
            table=bang, name=ten, code=ma, field_type=kieu, meaning=nhan, order=i,
        )
    ColumnDef.objects.create(
        table=bang, name="AOV", code="aov", field_type=FieldType.MONEY, order=5,
        is_computed=True, compute_op=ComputeOp.DIVIDE,
        compute_left="doanh_so", compute_right="so_don", compute_decimals=2,
    )
    return bang


@pytest.fixture
def bang_mkt(departments, nguoi_dung):
    """Bảng của bộ phận Marketing — kiểm chiều bị từ chối."""
    bang = TableDef.objects.create(
        name="Báo cáo MKT", code="bc_mkt_scope",
        department=departments["mkt"], created_by=nguoi_dung["manager_mkt"],
    )
    ColumnDef.objects.create(
        table=bang, name="Doanh số", code="doanh_so",
        field_type=FieldType.MONEY, meaning=Meaning.REVENUE, order=0,
    )
    return bang


#: (người tạo, ngày, người bán, sản phẩm, số đơn, doanh số)
DONG = [
    ("staff_sale_1", "2026-08-01", "Dũng", "SP A", "2", "1000"),
    ("staff_sale_1", "2026-08-02", "Dũng", "SP B", "1", "500"),
    ("staff_sale_1b", "2026-08-01", "Hà", "SP A", "1", "700"),
    ("staff_sale_2", "2026-08-01", "Hằng", "SP A", "3", "900"),
]


@pytest.fixture
def dong_sale(bang_bc_sale, nguoi_dung):
    """Bốn dòng: hai của staff_sale_1, một của người cùng team, một của team
    khác — đủ để soi từng cấp bậc."""
    for nguoi, ngay, ban, sp, don, ds in DONG:
        record_service.create_record(
            bang_bc_sale,
            {"ngay": ngay, "nguoi_ban": ban, "san_pham": sp,
             "so_don": don, "doanh_so": ds},
            actor=nguoi_dung[nguoi],
        )
    return bang_bc_sale


def _xem(client, nguoi_dung, ma, **tham_so):
    client.force_login(nguoi_dung[ma])
    tham_so.setdefault("tu", "2026-08-01")
    tham_so.setdefault("den", "2026-08-31")
    duoi = "&".join(f"{k}={v}" for k, v in tham_so.items())
    return client.get(f"{DUONG_DAN}?{duoi}")


# ══ Phạm vi quyền — FR-5.5 ═════════════════════════════════════════

def test_staff_chi_thay_tong_cua_minh(client, dong_sale, nguoi_dung):
    """AC-3.1 — Staff chỉ thấy bản ghi do chính mình tạo, không thấy của
    người cùng team"""
    phan_hoi = _xem(client, nguoi_dung, "staff_sale_1", nhom="nhan-vien")
    noi_dung = phan_hoi.content.decode()
    assert phan_hoi.status_code == 200
    # 1.000 + 500 của chính mình; 700 của người cùng team không được lẫn vào
    assert "1.500" in noi_dung
    assert "700" not in noi_dung
    assert "Hà" not in noi_dung


def test_leader_chi_thay_so_lieu_team_minh(client, dong_sale, nguoi_dung):
    """AC-5.5 — Leader chỉ thấy số liệu của team mình trong báo cáo tổng hợp

    Kiểm cả hai chiều: tổng đúng bằng team mình (1.000 + 500 + 700), và số
    của team khác (900, người bán Hằng) không xuất hiện.
    """
    phan_hoi = _xem(client, nguoi_dung, "leader_sale_1", nhom="nhan-vien")
    noi_dung = phan_hoi.content.decode()
    assert "2.200" in noi_dung
    assert "900" not in noi_dung
    assert "Hằng" not in noi_dung


def test_manager_thay_ca_bo_phan(client, dong_sale, nguoi_dung):
    """AC-3.4 — Manager thấy toàn bộ bản ghi của bộ phận mình"""
    phan_hoi = _xem(client, nguoi_dung, "manager_sale", nhom="nhan-vien")
    noi_dung = phan_hoi.content.decode()
    assert "3.100" in noi_dung
    assert "Hằng" in noi_dung


def test_admin_thay_moi_bang_trong_o_nguon(client, dong_sale, bang_mkt, nguoi_dung):
    """AC-3.8 — Quản trị viên thấy bảng của mọi bộ phận trong ô Nguồn số liệu"""
    phan_hoi = _xem(client, nguoi_dung, "admin", nguon="bc_sale")
    noi_dung = phan_hoi.content.decode()
    assert "bc_sale" in noi_dung
    assert "bc_mkt_scope" in noi_dung


def test_nguon_ngoai_pham_vi_tra_403(client, dong_sale, bang_mkt, nguoi_dung):
    """AC-3.6 — Truy cập dữ liệu ngoài phạm vi trả về lỗi từ chối, không
    phải danh sách rỗng

    Bảng của bộ phận khác và mã bảng không tồn tại đều 403 như nhau — không
    để lộ bảng nào có thật. Kiểm cả màn hình lẫn đường xuất Excel.
    """
    assert _xem(client, nguoi_dung, "staff_sale_1", nguon="bc_mkt_scope").status_code == 403
    assert _xem(client, nguoi_dung, "staff_sale_1", nguon="khong_ton_tai").status_code == 403
    phan_hoi = client.get(f"{DUONG_XUAT}?nguon=bc_mkt_scope")
    assert phan_hoi.status_code == 403
    # Bảng bộ phận khác cũng không hiện trong ô chọn
    noi_dung = _xem(client, nguoi_dung, "staff_sale_1").content.decode()
    assert "bc_mkt_scope" not in noi_dung


def test_nguoi_khong_co_bang_nguon_thay_trang_thai_rong(client, nguoi_dung):
    """NFR-6 — Chưa có bảng nguồn nào thì hiện trạng thái rỗng, không lỗi 500"""
    phan_hoi = _xem(client, nguoi_dung, "staff_vd")
    assert phan_hoi.status_code == 200
    assert "Chưa có bảng nguồn nào" in phan_hoi.content.decode()


# ══ Bộ lọc và dòng tổng cộng — FR-5.2 tới FR-5.4 ═══════════════════

def test_loc_khoang_thoi_gian_dung_so_ban_ghi(client, dong_sale, nguoi_dung):
    """AC-5.2 — Lọc theo khoảng thời gian trả về đúng số bản ghi trong
    khoảng đó"""
    phan_hoi = _xem(client, nguoi_dung, "manager_sale",
                    tu="2026-08-02", den="2026-08-02")
    noi_dung = phan_hoi.content.decode()
    # Chỉ một bản ghi ngày 02.08 — tổng cộng 1 ngày, doanh số 500
    assert "Tổng cộng · 1 ngày" in noi_dung
    assert "500" in noi_dung
    assert "1.000" not in noi_dung


def test_loc_san_pham_dung_so_ban_ghi(client, dong_sale, nguoi_dung):
    """AC-5.3 — Lọc theo sản phẩm trả về đúng số bản ghi"""
    phan_hoi = _xem(client, nguoi_dung, "manager_sale", nhom="nhan-vien",
                    sp="SP+B")
    noi_dung = phan_hoi.content.decode()
    assert "Tổng cộng · 1 nhân viên" in noi_dung
    assert "500" in noi_dung
    assert "2.600" not in noi_dung and "3.100" not in noi_dung


def test_dong_tong_cong_bang_tong_dong_chi_tiet(client, dong_sale, nguoi_dung):
    """AC-5.4 — Dòng tổng cộng bằng đúng tổng các dòng chi tiết

    So bằng Decimal trên dữ liệu trả về của tầng phép tính, và soi luôn cột
    tính sẵn: AOV tổng = tổng doanh số ÷ tổng số đơn.
    """
    from forms_builder.models import DataRecord
    from reports import aggregations

    kq = aggregations.summarize(
        dong_sale, DataRecord.objects.in_scope(nguoi_dung["manager_sale"]),
        group_key="nhan-vien",
    )
    dong = list(kq.rows)
    assert kq.totals["c_doanh_so"] == sum(
        (d["c_doanh_so"] for d in dong), Decimal("0"))
    assert kq.totals["c_doanh_so"] == Decimal("3100")
    assert kq.totals["aov"] == (Decimal("3100") / Decimal("7")).quantize(Decimal("0.01"))


def test_tab_thi_truong_hien_ghi_chu_hoan(client, dong_sale, nguoi_dung):
    """FR-5.1 — Tab thị trường hoãn theo Q36: hiện ghi chú, không có bảng số

    Cách nhóm thứ tư chưa chốt nguồn số liệu (backlog N9) nên tab chỉ có khối
    giải thích; nút Xuất Excel cũng ẩn đi.
    """
    phan_hoi = _xem(client, nguoi_dung, "manager_sale", nhom="thi-truong")
    noi_dung = phan_hoi.content.decode()
    assert phan_hoi.status_code == 200
    assert "bao-cho" in noi_dung
    assert "Chưa chốt nguồn số liệu" in noi_dung
    assert "<tbody" not in noi_dung
    assert "Xuất Excel" not in noi_dung


def test_man_hinh_khong_qua_muoi_lenh_truy_van(
        client, dong_sale, nguoi_dung, django_assert_max_num_queries):
    """AC-10.2 — Màn hình báo cáo tổng hợp chạy không quá 10 lệnh truy vấn

    Đo với Leader — vai nặng nhất: phạm vi phải đọc thêm team phụ trách.
    """
    client.force_login(nguoi_dung["leader_sale_1"])
    client.get(DUONG_DAN)                    # lượt đầu ghi mốc phiên
    with django_assert_max_num_queries(10):
        assert client.get(DUONG_DAN).status_code == 200


# ══ Xuất Excel — FR-5.6, nguyên tắc P5 ═════════════════════════════

def _tai_excel(client, nguoi_dung, ma, **tham_so):
    client.force_login(nguoi_dung[ma])
    tham_so.setdefault("tu", "2026-08-01")
    tham_so.setdefault("den", "2026-08-31")
    duoi = "&".join(f"{k}={v}" for k, v in tham_so.items())
    return client.get(f"{DUONG_XUAT}?{duoi}")


def test_tep_xuat_chua_dung_so_lieu(client, dong_sale, nguoi_dung):
    """FR-5.6 — Tệp xuất ra chứa đúng số liệu đang hiển thị trên màn hình

    AC-5.6 là tiêu chí thủ công (mở bằng Excel thật); bài này kiểm phần máy
    kiểm được: đọc lại tệp bằng openpyxl và đối chiếu từng số bằng Decimal.
    """
    phan_hoi = _tai_excel(client, nguoi_dung, "manager_sale", nhom="nhan-vien")
    assert phan_hoi.status_code == 200
    assert "spreadsheetml" in phan_hoi["Content-Type"]
    assert 'attachment; filename="bao-cao-tong-hop-nhan-vien-' in phan_hoi["Content-Disposition"]

    wb = load_workbook(BytesIO(phan_hoi.content))
    ws = wb.active
    cac_dong = list(ws.iter_rows(values_only=True))
    # Dòng 4 là tiêu đề cột, tiếp theo là dữ liệu, dòng cuối là tổng cộng
    assert cac_dong[3][0] == "Người bán"
    dong_cuoi = cac_dong[-1]
    assert dong_cuoi[0] == "Tổng cộng · 3 nhân viên"
    # Cột: Người bán, Số đơn, Doanh số, AOV
    assert Decimal(str(dong_cuoi[1])) == Decimal("7")
    assert Decimal(str(dong_cuoi[2])) == Decimal("3100")
    assert Decimal(str(dong_cuoi[3])) == (
        Decimal("3100") / Decimal("7")).quantize(Decimal("0.01"))
    # Ba dòng chi tiết cộng lại bằng dòng tổng — AC-5.4 trên chính tệp xuất
    chi_tiet = cac_dong[4:-1]
    assert sum((Decimal(str(d[2])) for d in chi_tiet), Decimal("0")) == Decimal("3100")


def test_leader_xuat_chi_co_so_team_minh(client, dong_sale, nguoi_dung):
    """AC-5.5 — Tệp Leader xuất ra chỉ chứa số liệu team mình"""
    phan_hoi = _tai_excel(client, nguoi_dung, "leader_sale_1", nhom="nhan-vien")
    wb = load_workbook(BytesIO(phan_hoi.content))
    cac_o = [str(o) for dong in wb.active.iter_rows(values_only=True) for o in dong]
    assert not any("Hằng" in o for o in cac_o)
    dong_cuoi = list(wb.active.iter_rows(values_only=True))[-1]
    assert Decimal(str(dong_cuoi[2])) == Decimal("2200")


def test_moi_lan_xuat_ghi_mot_dong_nhat_ky(client, dong_sale, nguoi_dung):
    """BR-6 — Mỗi lần xuất Excel ghi một dòng nhật ký, nguyên tắc P5

    Chi tiết chỉ ghi tham số lọc, không ghi số liệu (điều cấm 6).
    """
    truoc = AuditLog.objects.filter(action=AuditAction.EXPORT).count()
    _tai_excel(client, nguoi_dung, "manager_sale")
    ds = AuditLog.objects.filter(action=AuditAction.EXPORT)
    assert ds.count() == truoc + 1
    dong = ds.latest("created_at")
    assert "bc_sale" in dong.detail
    assert "3100" not in dong.detail


def test_xuat_tab_thi_truong_bi_chan(client, dong_sale, nguoi_dung):
    """FR-5.1 — Đường xuất từ chối tab thị trường đang hoãn, chuyển về màn
    hình chính kèm thông báo"""
    phan_hoi = _tai_excel(client, nguoi_dung, "manager_sale", nhom="thi-truong")
    assert phan_hoi.status_code == 302
    assert phan_hoi["Location"].startswith(DUONG_DAN)
