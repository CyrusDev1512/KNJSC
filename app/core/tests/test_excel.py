"""Kiểm đơn vị `core/excel.py`: nhận biết loại tệp, dò tiêu đề, ép kiểu ô, ghi bảng.

Không cần cơ sở dữ liệu. Mọi tệp Excel sinh trong bộ nhớ bằng openpyxl —
không commit tệp nhị phân làm fixture.
"""
import io
import unicodedata
from datetime import date, datetime
from decimal import Decimal

import pytest
from openpyxl import Workbook, load_workbook

from core import excel
from core.constants import UPLOAD_MAX_BYTES, FileKind


def _xlsx(rows):
    wb = Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    dem = io.BytesIO()
    wb.save(dem)
    dem.seek(0)
    return dem


def test_nhan_ra_excel_that():
    """AC-7.9 — Tệp .xlsx thật được nhận là Excel"""
    tep = _xlsx([["a", "b"]])
    assert excel.sniff_kind(tep, declared_name="don.xlsx") == FileKind.XLSX


def test_tu_choi_tep_doi_duoi():
    """AC-7.9 — Tệp không đúng định dạng cho phép bị từ chối

    Tệp .exe đổi đuôi .xlsx, và ảnh PNG đổi đuôi .csv — cả hai bị chặn vì
    đầu tệp không khớp đuôi khai báo (docs/03 S7).
    """
    exe = io.BytesIO(b"MZ\x90\x00" + b"\x00" * 100)
    with pytest.raises(excel.UploadRejected):
        excel.sniff_kind(exe, declared_name="virus.xlsx")

    png = io.BytesIO(b"\x89PNG\r\n\x1a\n" + b"\x00" * 50)
    with pytest.raises(excel.UploadRejected) as loi:
        excel.sniff_kind(png, declared_name="anh.csv")
    assert "đổi đuôi" in str(loi.value)


def test_excel_khai_la_csv_bi_tu_choi():
    """AC-7.9 — Nội dung Excel nhưng đuôi .csv cũng là tệp đổi đuôi"""
    with pytest.raises(excel.UploadRejected):
        excel.sniff_kind(_xlsx([["a"]]), declared_name="don.csv")


def test_csv_nhan_khi_doc_duoc_dang_chu():
    """NFR-12 — CSV không có chữ ký, nhận khi là chữ UTF-8 không byte 0"""
    tep = io.BytesIO("ten;so\nA;1\n".encode("utf-8-sig"))
    assert excel.sniff_kind(tep, declared_name="don.csv") == FileKind.CSV
    with pytest.raises(excel.UploadRejected):
        excel.sniff_kind(io.BytesIO(b"ten\x00so"), declared_name="don.csv")


def test_duoi_la_bi_tu_choi():
    """NFR-12 — Đuôi ngoài danh sách cho phép thì từ chối dù nội dung là gì"""
    with pytest.raises(excel.UploadRejected) as loi:
        excel.sniff_kind(_xlsx([["a"]]), declared_name="don.docx")
    assert "không được phép" in str(loi.value)


def test_thu_hep_theo_luong():
    """FR-7.5 — Luồng nhập chỉ nhận Excel và CSV, ảnh bị từ chối"""
    png = io.BytesIO(b"\x89PNG\r\n\x1a\n" + b"\x00" * 50)
    with pytest.raises(excel.UploadRejected):
        excel.sniff_kind(png, declared_name="anh.png", allowed=(FileKind.XLSX, FileKind.CSV))


def test_kich_thuoc_qua_10mb():
    """AC-7.8 — Tệp vượt 10 MB bị từ chối với thông báo rõ ràng"""
    excel.check_size(UPLOAD_MAX_BYTES)          # đúng ngưỡng thì qua
    with pytest.raises(excel.UploadRejected) as loi:
        excel.check_size(UPLOAD_MAX_BYTES + 1)
    assert "10,0 MB" in str(loi.value) and "vượt giới hạn" in str(loi.value)


def test_do_hang_tieu_de_khong_o_hang_dau():
    """FR-7.5 — Tệp thật để tiêu đề ở hàng 2 kèm ô xuống dòng, hàng 3 công thức"""
    rows = [
        [None, None, None],
        ["\nLọc trùng", "Name", "Phone"],
        [None, None, 221],
        [1, "Taylor Minh", "559-7393026"],
    ]
    chi_so, diem = excel.find_header_row(rows, ["Name", "Phone", "Tên khách"])
    assert (chi_so, diem) == (1, 2)
    assert excel.find_header_row(rows, ["Không có"]) == (0, 0)


def test_chuan_hoa_ten_cot():
    """FR-7.5 — Tên cột so khớp không phân biệt hoa thường, khoảng trắng, dạng mã Unicode"""
    assert excel.normalise_label("\nLọc trùng") == excel.normalise_label("lọc  TRÙNG ")
    nfd = unicodedata.normalize("NFD", "Hóa đơn")
    assert excel.normalise_label(nfd) == excel.normalise_label("hóa đơn")
    assert excel.normalise_label(None) == ""


def test_ep_kieu_o_excel():
    """FR-7.5 — Điện thoại dạng số thực và thời gian dạng chuỗi trong tệp thật đọc đúng"""
    assert excel.coerce_cell(7788599010.0) == 7788599010
    assert excel.coerce_cell(218.0) == 218
    assert excel.coerce_cell(1234.567) == 1234.567
    assert excel.coerce_cell("0:58 14/10/2023") == date(2023, 10, 14)
    assert excel.coerce_cell("00:53:00 16/10/2023") == date(2023, 10, 16)
    assert excel.coerce_cell("14/10/2023") == date(2023, 10, 14)
    assert excel.coerce_cell("31/02/2023") == "31/02/2023"      # ngày không có thật: giữ chữ
    assert excel.coerce_cell("  Calgary ") == "Calgary"
    assert excel.coerce_cell("   ") is None
    moc = datetime(2023, 11, 8)
    assert excel.coerce_cell(moc) is moc
    assert excel.coerce_cell(True) is True


def test_doc_bang_excel_bo_hang_trong_cuoi():
    """FR-7.5 — Hàng trống ở cuối sheet không tính là dòng"""
    tep = _xlsx([["a", "b"], [1, 2], [None, None], [None, None]])
    du_lieu = excel.read_table(tep, FileKind.XLSX)
    assert du_lieu.rows == [["a", "b"], [1, 2]]
    assert du_lieu.truncated is False


def test_doc_bang_cat_o_gioi_han():
    """NFR-13 — Đọc quá `max_rows` thì đánh dấu bị cắt để tầng trên từ chối"""
    tep = _xlsx([["a"]] + [[i] for i in range(30)])
    du_lieu = excel.read_table(tep, FileKind.XLSX, max_rows=5)
    assert du_lieu.truncated is True


def test_doc_csv_dau_cham_phay_va_bom():
    """FR-7.5 — CSV kiểu Việt Nam: dấu chấm phẩy, có BOM"""
    tep = io.BytesIO("ten;so\nA;1\nB;2\n".encode("utf-8-sig"))
    du_lieu = excel.read_table(tep, FileKind.CSV)
    assert du_lieu.rows == [["ten", "so"], ["A", "1"], ["B", "2"]]


def test_excel_hong_khong_no():
    """NFR-6 — Tệp có đầu zip nhưng không phải Excel → thông báo tiếng Việt, không 500"""
    gia = io.BytesIO(b"PK\x03\x04" + b"rac" * 100)
    with pytest.raises(excel.UploadRejected) as loi:
        excel.read_table(gia, FileKind.XLSX)
    assert "Không mở được tệp Excel" in str(loi.value)


def test_ghi_bang_giu_decimal_va_ngay():
    """BR-8 — Tệp xuất giữ Decimal nguyên trạng, ngày là ngày thật, tiêu đề đậm"""
    wb = excel.write_table(
        ["Ngày", "Tiền"],
        [[date(2026, 8, 1), Decimal("1234.56")], [date(2026, 8, 2), None]],
        sheet_title="Thu",
    )
    dem = io.BytesIO()
    wb.save(dem)
    dem.seek(0)
    ws = load_workbook(dem).active
    assert ws.title == "Thu"
    assert ws["A1"].font.b is True
    assert ws.freeze_panes == "A2"
    assert ws["A2"].value == datetime(2026, 8, 1)
    assert Decimal(str(ws["B2"].value)) == Decimal("1234.56")
    assert ws["B3"].value is None
