"""Đọc và ghi tệp Excel, CSV — Giai đoạn 7.

Chỉ chứa cơ khí tệp: nhận biết loại tệp, đọc bảng thô, dò hàng tiêu đề, ép
kiểu ô, ghi bảng. Không truy vấn, không biết bảng động là gì — tầng dịch vụ
(`import_service`, `export_service`) mới ghép hai thứ lại.

Ba nguyên tắc:

- **Kiểm loại tệp bằng đầu tệp, không tin đuôi** (docs/03 S7). `.xlsx` là
  một tệp zip nên bắt đầu bằng ``PK\\x03\\x04``; ảnh có chữ ký riêng; CSV không
  có chữ ký nên chỉ nhận khi đọc được dạng chữ và không có byte 0.
- **Tự dò hàng tiêu đề.** Tệp thật của công ty để tiêu đề ở hàng 2 (ô đầu
  còn có xuống dòng), hàng 3 là công thức đếm, dữ liệu từ hàng 4. Giả định
  tiêu đề ở hàng 1 là nhập sai từ đầu.
- **Số ghi ra là `Decimal` nguyên trạng**, không đổi sang float (BR-8) —
  theo đúng khuôn `reports/excel.py`.
"""
import csv
import io
import re
import unicodedata
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils.exceptions import InvalidFileException

from .constants import (
    FILE_EXTENSIONS, FILE_MAGIC, HEADER_SCAN_ROWS, UPLOAD_MAX_BYTES, FileKind,
)
from .exceptions import BusinessError


class UploadRejected(BusinessError):
    """Tệp tải lên bị từ chối — thông báo tiếng Việt nói rõ vì sao (NFR-6)."""

    code = "upload_rejected"


# ══ NHẬN BIẾT TỆP ═════════════════════════════════════════════════

def format_size(so_byte):
    """12.582.912 → "12,0 MB" — để thông báo đọc được."""
    mb = Decimal(so_byte) / Decimal(1024 * 1024)
    text = f"{mb:.1f}".replace(".", ",")
    return f"{text} MB"


def check_size(so_byte):
    """Từ chối tệp quá 10 MB với thông báo rõ ràng — NFR-11, AC-7.8."""
    if so_byte > UPLOAD_MAX_BYTES:
        raise UploadRejected(
            f"Tệp {format_size(so_byte)} vượt giới hạn {format_size(UPLOAD_MAX_BYTES)}. "
            "Chia nhỏ tệp rồi tải lên lại."
        )


def _dau_tep(upload, n=16):
    """Đọc n byte đầu của một đối tượng tệp rồi trả con trỏ về đầu."""
    upload.seek(0)
    dau = upload.read(n)
    upload.seek(0)
    return dau


def _giong_csv(upload):
    """CSV không có chữ ký: đọc được dạng chữ và không chứa byte 0 thì nhận."""
    upload.seek(0)
    mau = upload.read(4096)
    upload.seek(0)
    if not mau or b"\x00" in mau:
        return False
    try:
        mau.decode("utf-8-sig")
    except UnicodeDecodeError:
        return False
    return True


def kind_of_extension(ten_tep):
    """Loại tệp theo đuôi khai báo. Đuôi lạ trả None."""
    duoi = Path(str(ten_tep or "")).suffix.lower()
    for loai, cac_duoi in FILE_EXTENSIONS.items():
        if duoi in cac_duoi:
            return loai
    return None


def sniff_kind(upload, *, declared_name=None, allowed=None):
    """Loại tệp **thật** của một tệp tải lên — S7, AC-7.9.

    `declared_name` là tên tệp người dùng gửi: đuôi khai một đằng mà nội dung
    một nẻo thì từ chối, kể cả khi nội dung tự nó hợp lệ (tệp .exe đổi đuôi
    .xlsx là trường hợp kinh điển). `allowed` thu hẹp thêm — ví dụ luồng nhập
    chỉ nhận Excel và CSV.
    """
    dau = _dau_tep(upload)
    loai = None
    for ung_vien, cac_chu_ky in FILE_MAGIC.items():
        if any(dau.startswith(ck) for ck in cac_chu_ky):
            loai = ung_vien
            break
    if loai is None and _giong_csv(upload):
        loai = FileKind.CSV

    if loai is None:
        raise UploadRejected(
            "Không nhận ra loại tệp. Chỉ nhận Excel (.xlsx), CSV, ảnh JPG hoặc PNG."
        )

    if declared_name:
        theo_duoi = kind_of_extension(declared_name)
        if theo_duoi is None:
            raise UploadRejected(
                f'Đuôi tệp "{Path(str(declared_name)).suffix}" không được phép. '
                "Chỉ nhận .xlsx, .csv, .jpg, .png."
            )
        if theo_duoi != loai:
            raise UploadRejected(
                f"Tệp khai là {FileKind(theo_duoi).label} nhưng nội dung là "
                f"{FileKind(loai).label}. Không nhận tệp đổi đuôi."
            )

    if allowed is not None and loai not in allowed:
        raise UploadRejected(
            f"Chỗ này chỉ nhận {' hoặc '.join(FileKind(k).label for k in allowed)}."
        )
    return loai


# ══ ĐỌC BẢNG THÔ ══════════════════════════════════════════════════

@dataclass
class SheetData:
    """Bảng thô đọc từ tệp: các hàng còn nguyên giá trị, chưa hiểu cột nào là gì."""

    sheet_name: str
    rows: list = field(default_factory=list)   # mỗi hàng là list giá trị
    #: Số hàng trong tệp nhiều hơn số đã đọc (bị cắt ở `max_rows`)
    truncated: bool = False

    @property
    def row_count(self):
        return len(self.rows)


def _hang_trong(hang):
    return all(v in (None, "") for v in hang)


def _cat_duoi_rong(rows):
    """Bỏ các hàng trống ở cuối — Excel hay có hàng rỗng tới tận 2.099."""
    while rows and _hang_trong(rows[-1]):
        rows.pop()
    return rows


def read_table(nguon, kind, *, max_rows=None):
    """Đọc sheet đầu tiên của tệp thành `SheetData`.

    `nguon` là đường dẫn hoặc đối tượng tệp nhị phân. `max_rows` cắt số hàng
    đọc (đã trừ hàng trống cuối); vượt thì đặt `truncated` để tầng trên từ
    chối trước khi làm gì nặng (NFR-13).
    """
    if kind == FileKind.XLSX:
        return _doc_xlsx(nguon, max_rows)
    if kind == FileKind.CSV:
        return _doc_csv(nguon, max_rows)
    raise UploadRejected("Chỉ đọc được bảng từ tệp Excel hoặc CSV.")


def _gioi_han(rows, max_rows):
    rows = _cat_duoi_rong(rows)
    if max_rows is not None and len(rows) > max_rows:
        return rows[:max_rows], True
    return rows, False


def _doc_xlsx(nguon, max_rows):
    try:
        wb = load_workbook(nguon, read_only=True, data_only=True)
    except (InvalidFileException, zipfile.BadZipFile, KeyError, OSError, ValueError) as loi:
        raise UploadRejected(
            "Không mở được tệp Excel — tệp hỏng hoặc không phải định dạng .xlsx."
        ) from loi
    try:
        ws = wb.worksheets[0]
        # Đọc dư một khoảng cho phần dò tiêu đề, rồi mới cắt
        tran = None if max_rows is None else max_rows + HEADER_SCAN_ROWS + 1
        rows = []
        for i, hang in enumerate(ws.iter_rows(values_only=True)):
            rows.append(list(hang))
            if tran is not None and i + 1 >= tran:
                break
        ten = ws.title
    finally:
        wb.close()
    rows, cat = _gioi_han(rows, None if max_rows is None else max_rows + HEADER_SCAN_ROWS)
    return SheetData(sheet_name=ten, rows=rows, truncated=cat)


def _doc_csv(nguon, max_rows):
    if hasattr(nguon, "read"):
        nguon.seek(0)
        noi_dung = nguon.read()
    else:
        noi_dung = Path(nguon).read_bytes()
    try:
        text = noi_dung.decode("utf-8-sig")
    except UnicodeDecodeError as loi:
        raise UploadRejected("Tệp CSV phải lưu bằng mã UTF-8.") from loi
    mau = text[:4096]
    try:
        phan_cach = csv.Sniffer().sniff(mau, delimiters=",;\t").delimiter
    except csv.Error:
        phan_cach = ","
    rows = [
        [_o_csv(o) for o in hang]
        for hang in csv.reader(io.StringIO(text), delimiter=phan_cach)
    ]
    rows, cat = _gioi_han(rows, None if max_rows is None else max_rows + HEADER_SCAN_ROWS)
    return SheetData(sheet_name="CSV", rows=rows, truncated=cat)


def _o_csv(o):
    """Ô CSV luôn là chuỗi; ô rỗng thành None cho đồng nhất với Excel."""
    o = o.strip()
    return o if o else None


# ══ TIÊU ĐỀ VÀ Ô ═════════════════════════════════════════════════

_KHOANG_TRANG = re.compile(r"\s+")


def normalise_label(text):
    """Chuẩn hoá tên cột để so khớp: NFC, gộp khoảng trắng và xuống dòng,
    không phân biệt hoa thường. "\\nLọc trùng" và "lọc  trùng" là một."""
    if text is None:
        return ""
    text = unicodedata.normalize("NFC", str(text))
    return _KHOANG_TRANG.sub(" ", text).strip().casefold()


def find_header_row(rows, expected_labels):
    """Tìm hàng tiêu đề: hàng khớp nhiều tên cột mong đợi nhất trong
    `HEADER_SCAN_ROWS` hàng đầu. Trả `(chỉ số hàng, số cột khớp)`; không hàng
    nào khớp thì `(0, 0)`."""
    mong_doi = {normalise_label(x) for x in expected_labels if x}
    tot_nhat, diem_tot = 0, 0
    for i, hang in enumerate(rows[:HEADER_SCAN_ROWS]):
        nhan = {normalise_label(v) for v in hang if isinstance(v, str)}
        diem = len(nhan & mong_doi)
        if diem > diem_tot:
            tot_nhat, diem_tot = i, diem
    return tot_nhat, diem_tot


#: Ngày kiểu Việt Nam, có hoặc không kèm giờ. Dấu gạch chéo cho phép gõ đúp
#: (`1/12//2023`) — tệp thật có hai dòng như vậy, và tệp thật phải nhập được
#: mà không chỉnh sửa (docs/04 mục 13 điều 5).
_NGAY_GIO = re.compile(
    r"^\s*(?:(\d{1,2}):(\d{2})(?::(\d{2}))?\s+)?(\d{1,2})/+(\d{1,2})/+(\d{4})"
    r"(?:\s+(\d{1,2}):(\d{2})(?::(\d{2}))?)?\s*$"
)


def coerce_cell(value):
    """Đưa một ô Excel về dạng tầng nhập hiểu được.

    - Số thực nguyên (7788599010.0) → số nguyên: điện thoại Excel hay tự đổi
      sang số, giữ ".0" là hỏng cột chữ.
    - Chuỗi ngày kiểu Việt Nam `14/10/2023`, có hoặc không kèm giờ
      (`0:58 14/10/2023`) → `date`.
    - `datetime` giữ nguyên — `parse_value` đã nhận.
    - Chuỗi thì cắt khoảng trắng hai đầu; rỗng thành None.
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return value
    if isinstance(value, (int, Decimal, date, datetime)):
        return value
    text = str(value).strip()
    if not text:
        return None
    m = _NGAY_GIO.match(text)
    if m:
        ngay, thang, nam = int(m.group(4)), int(m.group(5)), int(m.group(6))
        try:
            return date(nam, thang, ngay)
        except ValueError:
            return text
    return text


# ══ GHI BẢNG ══════════════════════════════════════════════════════

def write_table(headers, rows, *, sheet_title="Du lieu"):
    """Ghi một bảng thành Workbook: hàng tiêu đề đậm và cố định, số giữ
    `Decimal` nguyên trạng, ngày là ngày thật để Excel hiểu."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]
    ws.append(list(headers))
    for o in ws[1]:
        o.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for hang in rows:
        ws.append([_o_ghi(v) for v in hang])
    return wb


def _o_ghi(v):
    """Giá trị đưa vào ô Excel. Kiểu số, ngày, Decimal giữ nguyên; còn lại là chuỗi."""
    if v is None:
        return None
    if isinstance(v, (int, float, Decimal, date, datetime, bool)):
        return v
    return str(v)
