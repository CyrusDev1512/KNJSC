#!/usr/bin/env python3
"""Ẩn danh hoá tệp vận đơn thật để đưa vào kho làm bài kiểm "nhập tệp thật".

    python scripts/an-danh-vandon.py <tệp gốc> docs/tham-khao/vandon-mau.xlsx

Giữ nguyên mọi thứ làm nên "bẫy" của tệp thật — hàng tiêu đề ở hàng 2 với ô
xuống dòng, hàng 3 công thức, điện thoại dạng số thực, thời gian dạng chuỗi,
danh sách chọn ở sheet `data`, số dòng — chỉ thay bốn cột nhận dạng khách:
Name, Phone, Add, Tên người chuyển tiền. Dữ liệu giả sinh theo seed cố định
nên chạy lại ra cùng một tệp.

Bản gốc không vào git (đặt ở storage/ hoặc ngoài kho).
"""
import random
import sys
from pathlib import Path

from openpyxl import load_workbook

HO = ["Nguyen", "Tran", "Le", "Pham", "Hoang", "Vu", "Dang", "Bui", "Do", "Ngo",
      "Smith", "Johnson", "Brown", "Wilson", "Taylor", "Lee", "Martin", "Garcia"]
TEN = ["An", "Binh", "Chi", "Dung", "Giang", "Hanh", "Khanh", "Lan", "Minh", "Nga",
       "Emma", "Liam", "Olivia", "Noah", "Ava", "Mia", "Lucas", "Amelia"]
DUONG = ["Maple Ave", "King St", "Queen St W", "Yonge St", "Main St", "Elm Dr",
         "Pine Rd", "Cedar Cres", "Oak Blvd", "Lakeshore Rd"]


def gia(ho_ten_goc, rng):
    return f"{rng.choice(HO)} {rng.choice(TEN)}"


def an_danh(nguon, dich):
    rng = random.Random(20260903)
    wb = load_workbook(nguon)
    ws = wb["VẬN ĐƠN"]
    tieu_de = {ws.cell(2, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(2, c).value}
    cot_ten = tieu_de["Name"]
    cot_dt = tieu_de["Phone"]
    cot_dc = tieu_de["Add"]
    cot_ck = tieu_de["Tên người chuyển tiền"]
    da_thay = 0
    for r in range(4, ws.max_row + 1):
        o_ten = ws.cell(r, cot_ten)
        if o_ten.value in (None, ""):
            continue
        ten_moi = gia(o_ten.value, rng)
        o_ten.value = ten_moi
        # Giữ nguyên "kiểu" điện thoại: số thực thì vẫn số thực, chữ thì vẫn chữ
        o_dt = ws.cell(r, cot_dt)
        so = f"{rng.randint(400, 999)}{rng.randint(2000000, 9999999)}"
        if isinstance(o_dt.value, float):
            o_dt.value = float(so)
        elif isinstance(o_dt.value, int):
            o_dt.value = int(so)
        elif o_dt.value not in (None, ""):
            o_dt.value = f"{so[:3]}-{so[3:]}"
        o_dc = ws.cell(r, cot_dc)
        if o_dc.value not in (None, ""):
            o_dc.value = f"{rng.randint(1, 999)} {rng.choice(DUONG)}"
        o_ck = ws.cell(r, cot_ck)
        if o_ck.value not in (None, ""):
            o_ck.value = ten_moi if rng.random() < 0.7 else gia(o_ck.value, rng)
        da_thay += 1
    # Sheet tài khoản thanh toán chứa link/số tài khoản thật — xoá sạch nội dung
    if "Thông tin Tài khoản Thanh Toán" in wb.sheetnames:
        ws2 = wb["Thông tin Tài khoản Thanh Toán"]
        for row in ws2.iter_rows():
            for o in row:
                o.value = None
        ws2["A1"] = "Đã bỏ nội dung khi ẩn danh hoá"
    Path(dich).parent.mkdir(parents=True, exist_ok=True)
    wb.save(dich)
    return da_thay


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(2)
    n = an_danh(sys.argv[1], sys.argv[2])
    print(f"Đã ẩn danh {n} dòng → {sys.argv[2]}")
